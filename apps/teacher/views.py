# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

import openpyxl
import math #LIBRERIAS AGREGADAS PARA LA VALIDACION
import re #LIBRERIAS AGREGADAS PARA LA VALIDACION
import json
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
from rest_framework import generics

from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse, reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, TemplateView

from apps.teacher.forms import TeacherForm, SubjectForm
from apps.teacher.models import Teacher, Subject
from apps.teacher.serializers import TeacherSerializer,TeacherSerializer2, SubjectSerializer

from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, cm, letter,landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Paragraph, Table, TableStyle, Image, SimpleDocTemplate, PageBreak

# Create your views here.

#------ Vistas para Teachers

class TeacherList(ListView):
	#model = Teacher
	queryset = Teacher.objects.order_by('id_tea')
	template_name = 'teacher/teacher_list.html'

class TeacherCreate(CreateView):
	model = Teacher
	form_class = TeacherForm
	template_name = 'teacher/teacher_form.html'
	success_url = reverse_lazy('teacher:teacher_listar')

class TeacherUpdate(UpdateView):
	model = Teacher
	form_class = TeacherForm
	template_name = 'teacher/teacher_form.html'
	success_url = reverse_lazy('teacher:teacher_listar')

class TeacherDelete(DeleteView):
	model = Teacher
	template_name = 'teacher/teacher_delete.html'
	success_url = reverse_lazy('teacher:teacher_listar')

class TeacherAPI_List(generics.ListCreateAPIView):
	queryset = Teacher.objects.all().order_by('id_tea')
	serializer_class = TeacherSerializer

class TeacherAPI_Detail(generics.UpdateAPIView):
	queryset = Teacher.objects.all()
	serializer_class = TeacherSerializer
	#lookup_field = 'dni_tea'

def TeacherReport(request):
	response = HttpResponse(content_type='application/pdf')
	response['Content-Disposition'] = 'filename="Reporte_Docentes.pdf"'
	#response['Content-Disposition'] = 'attachment; filename="Report.pdf"'

	buffer = BytesIO()
	#c = canvas.Canvas(buffer, pagesize=A4
	c = canvas.Canvas(buffer, pagesize=A4)
	c.setTitle('Reporte de Docentes')
	

	#Header
	c.setLineWidth(.3)
	c.setFont('Helvetica', 22)
	c.drawString(4*cm, 27*cm,'SICOLAB')
	c.setFont('Helvetica', 14)
	c.drawString(4*cm, 26*cm,'Reporte de Carreras')
	img = Image('static/images/logo_carrera.jpg', width=2.5*cm, height=2.5*cm)
	img.drawOn(c, 17*cm, 25.6*cm)
	img = Image('static/images/sicolab_logo.png', width=3*cm, height=3*cm)
	img.drawOn(c, 0.8*cm, 25.2*cm)
	
	#Table header
	styles = getSampleStyleSheet()
	styleBH = styles['Normal']
	styleBH.alignment = TA_CENTER
	styleBH.fontSize = 10

	num = Paragraph('<b>#</b>', styleBH)
	ced = Paragraph('<b>Cédula</b>', styleBH)
	nom = Paragraph('<b>Docente</b>', styleBH)
	emai = Paragraph('<b>E-mail</b>', styleBH)
	est = Paragraph('<b>Estado</b>', styleBH)

	data = []
	data.append([num, ced, nom, emai, est])

	#Table
	styleN = styles['BodyText']
	styleN.alignment = TA_CENTER
	styleN.fontSize = 7
	high = 650
	
	for tea in Teacher.objects.all():
		aux=""
		if tea.status == 1:
			aux="Activo"
		else:
		  	aux="Inactivo"

		row = [tea['id_tea'], tea['dni_tea'], tea['names_tea'],tea['email_tea'], aux]
		data.append(row)
		high = high - 18
	
	#Table size
	width, height = A4
	table = Table(data, colWidths=[1*cm, 3*cm, 6.5*cm,6.5*cm, 2*cm])
	table.setStyle(TableStyle([
		('BACKGROUND', (0,0), (5,0), colors.lightblue),
		('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
		('BOX', (0,0), (-1,-1), 0.25, colors.black), ]))
	
	#pdf size
	table.wrapOn(c, width, height)
	table.drawOn(c, 30, high)
	c.showPage()

	c.save()
	pdf = buffer.getvalue()
	buffer.close()
	response.write(pdf)
	return response	

#------ Vistas para Subjects

class SubjectList(ListView):
	#model = Subject
	queryset = Subject.objects.order_by('id_sub')
	template_name = 'teacher/subject_list.html'

class SubjectCreate(CreateView):
	model = Subject
	form_class = SubjectForm
	template_name = 'teacher/subject_form.html'
	success_url = reverse_lazy('subject:subject_listar')

class SubjectUpdate(UpdateView):
	model = Subject
	form_class = SubjectForm
	template_name = 'teacher/subject_form.html'
	success_url = reverse_lazy('subject:subject_listar')

class SubjectDelete(DeleteView):
	model = Subject
	template_name = 'teacher/subject_delete.html'
	success_url = reverse_lazy('subject:subject_listar')

class SubjectAPI_List(generics.ListCreateAPIView):
	queryset = Subject.objects.all()
	serializer_class = SubjectSerializer

class SubjectAPI_Detail(generics.RetrieveUpdateDestroyAPIView):
	queryset = Subject.objects.all()
	serializer_class = SubjectSerializer

def SubjectReport(request):
	response = HttpResponse(content_type='application/pdf')
	response['Content-Disposition'] = 'filename="Reporte_Asignaturas.pdf"'
	#response['Content-Disposition'] = 'attachment; filename="Report.pdf"'

	buffer = BytesIO()
	#c = canvas.Canvas(buffer, pagesize=A4
	c = canvas.Canvas(buffer, pagesize=A4)
	c.setTitle('Reporte de Asignaturas')
	

	#Header
	def header():
		c.setLineWidth(.3)
		c.setFont('Helvetica', 22)
		c.drawString(4*cm, 27*cm,'SICOLAB')
		c.setFont('Helvetica', 14)
		c.drawString(4*cm, 26*cm,'Reporte de Asignaturas')
		img = Image('static/images/logo_carrera.jpg', width=2.5*cm, height=2.5*cm)
		img.drawOn(c, 17*cm, 25.6*cm)
		img = Image('static/images/sicolab_logo.png', width=3*cm, height=3*cm)
		img.drawOn(c, 0.8*cm, 25.2*cm)
	
	#Table header
	styles = getSampleStyleSheet()
	styleBH = styles['Normal']
	styleBH.alignment = TA_CENTER
	styleBH.fontSize = 10

	num = Paragraph('<b>#</b>', styleBH)
	asig = Paragraph('<b>Asignatura</b>', styleBH)


	data = []
	data.append([num, asig])

	#Table
	styleN = styles['BodyText']
	styleN.alignment = TA_CENTER
	styleN.fontSize = 7
	high = 650
	cont = 0
	num_pag = 0
	
	print 'Maximo:' + str(len(Subject.objects.all()))

	for sub in Subject.objects.all().order_by('id_sub'):		
		row = [sub['id_sub'], sub['name_sub']]
		data.append(row)
		high = high - 18		
		cont = cont +1
		#print 'Cont: ' + str(cont)
		if cont == 32:		
			cont = 0
			#print 'Cont if: ' + str(cont)
			num_pag = num_pag + 1

			header()

			#Table size
			width, height = A4
			table = Table(data, colWidths=[2*cm, 17*cm])
			table.setStyle(TableStyle([
				('BACKGROUND', (0,0), (5,0), colors.lightblue),
				('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
				('BOX', (0,0), (-1,-1), 0.25, colors.black), ]))
			
			#pdf size
			table.wrapOn(c, width, height)
			table.drawOn(c, 30, high)
			c.showPage()
			data = []
			data.append([num, asig])
			high = 23*cm

	#print 'contador: ' + str(cont)
	#print 'condicion: ' + str((cont + num_pag * 32))
	if (cont + num_pag * 32) == len(Subject.objects.all()):		
		cont = 0
		
		header()
	
		#Table size
		width, height = A4
		table = Table(data, colWidths=[2*cm, 17*cm])
		table.setStyle(TableStyle([
			('BACKGROUND', (0,0), (5,0), colors.lightblue),
			('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
			('BOX', (0,0), (-1,-1), 0.25, colors.black), ]))
		
		#pdf size
		table.wrapOn(c, width, height)
		table.drawOn(c, 30, high)
		c.showPage()
	
	#print 'tamaño: ' + str(len(Subject.objects.all()))

	c.save()
	pdf = buffer.getvalue()
	buffer.close()
	response.write(pdf)
	return response	


#--------CLASES Y FUNCIONES PARA IMPORTAR TEACHER

class TeacherImport(TemplateView):
	#model = UploadForm
	#form_class = UploadForm
	template_name = 'teacher/teacher_import.html'
	#success_url = reverse_lazy('computer:computer_listar')


def teacherCargarExcel(request):
	if request.method == 'POST':
		nombre_archivo = request.FILES.get('archivo')
		doc = openpyxl.load_workbook(nombre_archivo)
		nombres = doc.get_sheet_names()
		hoja = doc.get_sheet_by_name(nombres[0])
		datos = []
		lista = []
		error = ""
		cont = 0
		num = 2
		try:
			for filas in hoja.rows:
				if cont != 0:
					celda = "A"+repr(num)
					celdaE = "E"+repr(num)
					celdaD = "D"+repr(num)
					nro = hoja[celda].value
					nro2= hoja[celdaE].value
					nro3= hoja[celdaD].value
					
					if nro is None:
						break

					verificar_cedula(nro, celda)
					verificar_estado(nro2, celdaE)
					#validar_correo(nro3, celdaD)

					num = num + 1
				for columna in filas:
					if columna.value is not None:
						datos.append(columna.value)
					else:
						raise Exception('ERROR: existen celdas vacias')
				if len(datos)!=0:
						print datos
						lista.append(datos)
				datos = []
				cont = cont + 1
		except Exception as err:
			error = str(err)
		ruta = handle_uploaded_file(nombre_archivo)
		context = {'lista':lista, 'nombre_archivo': nombre_archivo, 'ruta':ruta, 'error': error}
		return render(request, 'teacher/teacher_list_excel.html', context)

def teacherCargarBase(request):
	if request.method == 'GET':
		ruta = request.GET['enlace']
		doc = openpyxl.load_workbook('media/' + ruta)
		nombres = doc.get_sheet_names()
		hoja = doc.get_sheet_by_name(nombres[0])
		datos = []
		lista = []
		for filas in hoja.rows:
			for columna in filas:
				if columna.value is not None:
					datos.append(columna.value)	
			if len(datos)!=0:
					print "Al cargar en base de datos:"
					print datos
					lista.append(datos)
			datos = []
		x = 0
		docentesRegistrados = Teacher.objects.all()
		print len(lista)
		for fila in lista:
			if x == 0:
				x = x + 1
			else:
				if Teacher.objects.filter(dni_tea=fila[0]).count() == 0:
					Teacher.objects.create(dni_tea=fila[0], password_tea=fila[1], names_tea=fila[2], email_tea=fila[3], status=fila[4])
		return HttpResponseRedirect(reverse('teacher:teacher_listar'))

		
def handle_uploaded_file(f):
    with open('media/' + str(f.name), 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return str(f.name)
 
#--------------------------------- VALIDACIONES :v -----------------------------

def verificar_cedula(nro, celda):
	l = len(nro)
	if l != 10:
		raise Exception('ERROR: en la celda '+celda+' existe una cédula incorrecta')

def verificar_estado(nro2, celdaE):
	if nro2 > 1:
		raise Exception('ERROR: en la celda '+celdaE+' existe un estado diferente de 1 o 0 (Activo o Inactivo)')

#--------CLASES Y FUNCIONES PARA IMPORTAR ASIGNATURAS :v 

class SubjectImport(TemplateView):
	#model = UploadForm
	#form_class = UploadForm
	template_name = 'teacher/subject_import.html'
	#success_url = reverse_lazy('computer:computer_listar')

def subjectCargarExcel(request):
	if request.method == 'POST':
		nombre_archivo = request.FILES.get('archivo')
		doc = openpyxl.load_workbook(nombre_archivo)
		nombres = doc.get_sheet_names()
		hoja = doc.get_sheet_by_name(nombres[0])
		datos = []
		lista = []
		error = ""
		cont = 0
		num = 2
	try:
		for filas in hoja.rows:
			for columna in filas:
				if columna.value is not None:
					datos.append(columna.value)
				else:
					raise Exception('ERROR: existen celdas vacias')
			if len(datos)!=0:
				print datos
				lista.append(datos)
			datos = []
			cont = cont + 1
	except Exception as err:
		error = str(err)
	ruta = handle_uploaded_file(nombre_archivo)
	context = {'lista':lista, 'nombre_archivo': nombre_archivo, 'ruta':ruta, 'error': error}
	return render(request, 'teacher/subject_list_excel.html', context)

def subjectCargarBase(request):
	if request.method == 'GET':
		ruta = request.GET['enlace']
		doc = openpyxl.load_workbook('media/' + ruta)
		nombres = doc.get_sheet_names()
		hoja = doc.get_sheet_by_name(nombres[0])
		datos = []
		lista = []
		for filas in hoja.rows:
			for columna in filas:
				if columna.value is not None:
					datos.append(columna.value)	
			if len(datos)!=0:
				print "Al cargar en base de datos:"
				print datos
				lista.append(datos)
			datos = []
		x = 0
		docentesRegistrados = Subject.objects.all()
		print len(lista)
		for fila in lista:
			if x == 0:
				x = x + 1
			else:
				if Subject.objects.filter(name_sub=fila[0]).count() == 0:
					Subject.objects.create(name_sub=fila[0])
		return HttpResponseRedirect(reverse('teacher:subject_listar'))
